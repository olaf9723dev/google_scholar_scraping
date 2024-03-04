import os
import time
import math
import random
import requests
import threading
import concurrent
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from gensim.models import Word2Vec
from nltk.tokenize import word_tokenize
from playwright_stealth import stealth_sync
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill, Alignment
from concurrent.futures import ThreadPoolExecutor

MAIN_URL = "https://scholar.google.com{}"
DATA_PATH = "config/testing_data.xlsx"
PROXIES = 'config/proxy_ips.xlsx'
USER_AGENTS = 'config/allowed_user_agents.txt'
SIMILARITY_RATE = 0.8

class Extractor:
    def __init__(self) -> None:
        self.keys = []
        self.proxies = []
        self.agents = []
        self.links = []
        self.user_counts = []      
        self.session = requests.Session()
        self.lock = threading.Lock()
        try:
            os.mkdir('output')
        except FileExistsError:
            pass

    def get_all(self):
        print('Start')
        if os.path.exists('output/result.xlsx'):
            os.remove('output/result.xlsx')
        else:
            pass
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'Email'
        ws.cell(row=1, column=2).value = 'Name'
        ws.cell(row=1, column=3).value = 'Result'
        ws.cell(row=1, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=1, column=4).value = 'Full Name'
        ws.cell(row=1, column=5).value = 'Domain'
        ws.cell(row=1, column=6).value = 'Expertise'
        ws.cell(row=1, column=7).value = 'cit_all'
        ws.cell(row=1, column=8).value = 'cit_since_2019'
        ws.cell(row=1, column=9).value = 'h_ind_all'
        ws.cell(row=1, column=10).value = 'h_ind_since_2019'
        ws.cell(row=1, column=11).value = 'i10_ind_all'
        ws.cell(row=1, column=12).value = 'i10_ind_since_2019'
        for i in range(1,12):
            ws.cell(row=1, column=i).alignment= Alignment(horizontal='center')
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 50
        wb.save('output/result.xlsx')
        self.read_keys(DATA_PATH)
        
        self.get_datas()
        
        print('Finished all')

    def get_data(self, key, index):
        try:
            print('NO:', index + 1, 'progress:', math.floor(index/len(self.keys)*1000)/1000, '%')
            query = 'https://scholar.google.com/scholar?hl=en&as_sdt=0%2C5&q={}%40{}%09{}&btnG='
            proxy = self.proxies[random.randint(0, len(self.proxies)-1)]
            print(proxy)
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    proxy={
                        'server': proxy['url'],
                    },
                    timeout=60000,
                )
                context = browser.new_context()
                page = context.new_page()
                stealth_sync(page)

                content = self.return_page(query.format(str(key['email']).split('@')[0], str(key['email']).split('@')[1], str(key['name']).replace(' ', '+')), page)
                soup_page = BeautifulSoup(content, 'html.parser')
                temp_link =''
                for element in soup_page.find(id = 'gs_res_ccl_mid').find_all('div', 'gs_r gs_or gs_scl'):
                    try:
                        for el in element.find('div', 'gs_a').find_all('a'):
                            leng = len(str(key['name']).split(' ')) 
                            if self.compare_closet(str(key['name']).split(' ')[leng - 1], str(el.get_text()).split(' ')[1]) > 0.8:
                                temp_link = el['href']
                                break;
                        if temp_link != "":
                            break
                    except:
                        pass
                if temp_link =='':
                    temp_data=dict()
                    temp_data['index'] = index + 2
                    temp_data['email'] = key['email']
                    temp_data['name'] = key['name']
                    temp_data['result'] = ''
                    # self.lock.acquire()
                    # self.save_result(temp_data)
                    # self.lock.release()
                    print("There is no link for this person : ", temp_data['name'])
                    browser.close()
                    return temp_data
                else:
                    temp_data=dict()
                    temp_data['index'] = index + 2
                    temp_data['email'] = key['email']
                    temp_data['name'] = key['name']
                    temp_data['result'] = MAIN_URL.format(temp_link)
                    temp_data = self.get_detail_data(temp_data, page)
                    # self.lock.acquire()
                    # self.save_result(temp_data)
                    # self.lock.release()
                    browser.close()
                    return temp_data
        except:
            pass
        pass

    def get_datas(self):
        datas = None
        data = dict()
        datas = []
        with ThreadPoolExecutor() as exe:
            futures = [exe.submit(self.get_data, key, index) for index, key in enumerate(self.keys)]
            
            for future in concurrent.futures.as_completed(futures):
                data = future.result()
                self.save_result(data)
                datas.append(data)
        self.save_results(datas)   
    
    def get_detail_data(self, temp_data, page):
        try:
            content = self.return_page(temp_data['result'], page)
            soup_page= BeautifulSoup(content, 'html.parser')
            temp_data['full_name'] = soup_page.find(id = "gsc_prf_inw").get_text()
            temp_data['domain'] = soup_page.find_all('div', 'gsc_prf_il')[1].get_text()
            temp_data['expertise'] = soup_page.find_all('div', 'gsc_prf_il')[2].get_text()
            temp_data['cit_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[0].find_all('td')[1].get_text()
            temp_data['cit_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[0].find_all('td')[2].get_text()
            temp_data['h_ind_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[1].find_all('td')[1].get_text()
            temp_data['h_ind_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[1].find_all('td')[2].get_text()
            temp_data['i10_ind_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[2].find_all('td')[1].get_text()
            temp_data['i10_ind_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[2].find_all('td')[2].get_text()
            return temp_data
        except Exception as e:
            print('Cant get detail data for ', temp_data['name'], ' : ', e)

    def save_result(self, temp_data):
        if temp_data != None: 
            workbook = load_workbook('output/result.xlsx')
            worksheet = workbook.active
            if temp_data['result']=='' or temp_data['result']== None:
                worksheet.cell(row=int(temp_data['index']), column=1).value = temp_data['email']
                worksheet.cell(row=int(temp_data['index']), column=2).value = temp_data['name']
                worksheet.cell(row=int(temp_data['index']), column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for i in range(1, 3):
                    worksheet.cell(row=int(temp_data['index']), column=i).alignment = Alignment(horizontal='center')
                workbook.save('output/result.xlsx')
            else:
                worksheet.cell(row=int(temp_data['index']), column=1).value = temp_data['email']
                worksheet.cell(row=int(temp_data['index']), column=2).value = temp_data['name']
                worksheet.cell(row=int(temp_data['index']), column=3).value = temp_data['result']
                worksheet.cell(row=int(temp_data['index']), column=4).value = temp_data['full_name']
                worksheet.cell(row=int(temp_data['index']), column=5).value = temp_data['domain']
                worksheet.cell(row=int(temp_data['index']), column=6).value = temp_data['expertise']
                worksheet.cell(row=int(temp_data['index']), column=7).value = temp_data['cit_all']
                worksheet.cell(row=int(temp_data['index']), column=8).value = temp_data['cit_since_2019']
                worksheet.cell(row=int(temp_data['index']), column=9).value = temp_data['h_ind_all']
                worksheet.cell(row=int(temp_data['index']), column=10).value = temp_data['h_ind_since_2019']
                worksheet.cell(row=int(temp_data['index']), column=11).value = temp_data['i10_ind_all']
                worksheet.cell(row=int(temp_data['index']), column=12).value = temp_data['i10_ind_since_2019']
                for i in range(1, 12):
                    worksheet.cell(row=int(temp_data['index']), column=i).alignment = Alignment(horizontal='center')
                workbook.save('output/result.xlsx')
        else:
            pass

    def save_results(self, temp_datas):
        for temp_data in temp_datas:
            if temp_data != None: 
                workbook = load_workbook('output/result.xlsx')
                worksheet = workbook.active
                if temp_data['result']=='' or temp_data['result']== None:
                    worksheet.cell(row=int(temp_data['index']), column=1).value = temp_data['email']
                    worksheet.cell(row=int(temp_data['index']), column=2).value = temp_data['name']
                    worksheet.cell(row=int(temp_data['index']), column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    for i in range(1, 3):
                        worksheet.cell(row=int(temp_data['index']), column=i).alignment = Alignment(horizontal='center')
                    workbook.save('output/result.xlsx')
                else:
                    worksheet.cell(row=int(temp_data['index']), column=1).value = temp_data['email']
                    worksheet.cell(row=int(temp_data['index']), column=2).value = temp_data['name']
                    worksheet.cell(row=int(temp_data['index']), column=3).value = temp_data['result']
                    worksheet.cell(row=int(temp_data['index']), column=4).value = temp_data['full_name']
                    worksheet.cell(row=int(temp_data['index']), column=5).value = temp_data['domain']
                    worksheet.cell(row=int(temp_data['index']), column=6).value = temp_data['expertise']
                    worksheet.cell(row=int(temp_data['index']), column=7).value = temp_data['cit_all']
                    worksheet.cell(row=int(temp_data['index']), column=8).value = temp_data['cit_since_2019']
                    worksheet.cell(row=int(temp_data['index']), column=9).value = temp_data['h_ind_all']
                    worksheet.cell(row=int(temp_data['index']), column=10).value = temp_data['h_ind_since_2019']
                    worksheet.cell(row=int(temp_data['index']), column=11).value = temp_data['i10_ind_all']
                    worksheet.cell(row=int(temp_data['index']), column=12).value = temp_data['i10_ind_since_2019']
                    for i in range(1, 12):
                        worksheet.cell(row=int(temp_data['index']), column=i).alignment = Alignment(horizontal='center')
                    workbook.save('output/result.xlsx')
            else:
                pass
    
    def read_keys(self, filepath):
        print('Reading keywords for searching...')
        try:
            workbook = load_workbook(filename=filepath)
            worksheet = workbook['Sheet1']
            
            for row in worksheet.iter_rows(values_only=True):
                if row[0] == None:
                    break
                row_data=dict()
                row_data['email'] = row[0]
                row_data['name'] = row[1]
                self.keys.append(row_data)
            self.keys.pop(0)
        except Exception as e:
            print('There is error: ', e, 'reading keywords.') 

    def read_proxies(self, filepath):
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook['Sheet1']

            for row in worksheet.iter_rows(values_only=True):
                if row[0] == None:
                    break
                proxy_url = row[0]

                proxy_address = {'url':proxy_url}
                self.proxies.append(proxy_address)
            self.proxies.pop(0)
        except Exception as e:
            print('There is ',e,'reading proxies.') 

    def compare_closet(self, str1='', str2=''):
        tokens1 = word_tokenize(str1.lower())
        tokens2 = word_tokenize(str2.lower())
        sentences = [tokens1, tokens2]
        model = Word2Vec(sentences, min_count = 1)
        similarity = model.wv.n_similarity(tokens1, tokens2)
        
        if similarity > SIMILARITY_RATE:
            return similarity
        else:
            return 0
    
    def return_page(self, url, page):
        page.goto(url)
        content = page.content()
        return content        

def main():
    start_time = time.time()
    extracotr = Extractor()
    extracotr.read_proxies(PROXIES)
    extracotr.get_all()
    end_time = time.time()
    print('Elapsed Time is', math.floor((end_time-start_time)*100)/100, 's')

if __name__ == "__main__":
   main()