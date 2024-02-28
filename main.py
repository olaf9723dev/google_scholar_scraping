import os
import requests
import random
import csv
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from nltk.tokenize import word_tokenize
from gensim.models import Word2Vec
from playwright.sync_api import sync_playwright

MAIN_URL = "https://scholar.google.com{}"
DATA_PATH = "config/testing_data.xlsx"
PROXIES = 'config/proxy_ips.xlsx'
SIMILARITY_RATE = 0.8
class Extractor:
    def __init__(self) -> None:
        self.keys = []
        self.links = []
        self.session = requests.Session()
        self.proxies = []
        try:
            os.mkdir('output')
        except FileExistsError:
            pass


    def get_all(self):
        if os.path.exists('output/result.csv'):
            os.remove('output/result.csv')
        else:
            pass
        
        i = random.randint(0, len(self.proxies) - 1)
        proxy = self.proxies[i]

        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=True
            # proxy={
            # "server": proxy['url'],
            # "username": proxy['username'],
            # "password": proxy['pwd']
            # }
        )
        self.context = self.browser.new_context()
        self.page = self.context.new_page()
        self.read_keys(DATA_PATH)
        self.get_datas()
        self.browser.close()

    def get_datas(self):
        query = 'https://scholar.google.com/scholar?hl=en&as_sdt=0%2C5&q={}%40{}%09{}&btnG='
        for key in self.keys:
            response = self.get_response(query.format(str(key['email']).split('@')[0], str(key['email']).split('@')[1], str(key['name']).replace(' ', '+')))
            content = response.text
            soup_page= BeautifulSoup(content, 'html.parser')
            try:
                temp_link=''
                for element in soup_page.find(id = "gs_res_ccl_mid").find_all('div', 'gs_r gs_or gs_scl'):
                    try:
                        for el in element.find('div', 'gs_a').find_all('a'):
                            leng = len(str(key['name']).split(' ')) 
                            if self.compare_closet(str(key['name']).split(' ')[leng - 1], str(el.get_text()).split(' ')[1]) > 0.8:
                                temp_link = el['href']
                                break;
                        if temp_link != "":
                            break
                    except:
                        pass

                if temp_link =="":
                    print("There is no link for this person")
                else:
                    temp_data=dict()
                    temp_data['email'] = key['email']
                    temp_data['name'] = key['name']
                    temp_data['result'] = MAIN_URL.format(temp_link)
                    temp_data = self.get_detail_data(temp_data)
                    self.update_csv(temp_data)
            except Exception as e:
                print("There is no result for ", key,'.')
        print('Finished all.')

    def get_detail_data(self, temp_data):
        print('Reading Detail data for ', temp_data['name'],'.')
        try:
            # response = self.get_response(temp_data['result'])
            # content = response.text
            content = self.return_detailpage(temp_data['result'])
            soup_page= BeautifulSoup(content, 'html.parser')
            temp_data['full_name'] = soup_page.find(id = "gsc_prf_inw").get_text()
            temp_data['domain'] = soup_page.find_all('div', 'gsc_prf_il')[1].get_text()
            temp_data['expertise'] = soup_page.find_all('div', 'gsc_prf_il')[2].get_text()
            temp_data['cit_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[0].find_all('td')[1].get_text()
            temp_data['cit_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[0].find_all('td')[2].get_text()
            temp_data['h_ind_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[1].find_all('td')[1].get_text()
            temp_data['h_ind_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[1].find_all('td')[2].get_text()
            temp_data['i10_ind_all'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[2].find_all('td')[1].get_text()
            temp_data['i10_ind_since_2019'] = soup_page.find(id = "gsc_rsb_st").find('tbody').find_all('tr')[2].find_all('td')[2].get_text()
            return temp_data
        except Exception as e:
            print('Cant get detail data for ', temp_data['name'], ' : ', e)

    def update_csv(self, temp_data):
        print('Added the ', temp_data['full_name'],' to Excel File...')
        with open('output/result.csv', 'a', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=temp_data.keys())
            if csvfile.tell() == 0:
                writer.writeheader()
            writer.writerow(temp_data)

    def read_keys(self, filepath):
        print('Reading keywords for searching...')
        try:
            workbook = load_workbook(filename=filepath)
            worksheet = workbook['Sheet1']
            
            for row in worksheet.iter_rows(values_only=True):
                if row[0] == None:
                    break
                row_data=dict()
                row_data['email'] = row[0]
                row_data['name'] = row[1]
                self.keys.append(row_data)
            self.keys.pop(0)
        except Exception as e:
            print('There is error: ', e, 'reading keywords.') 

    def read_proxies(self, filepath):
        print('Reading proxies ...')
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook['Sheet1']

            for row in worksheet.iter_rows(values_only=True):
                if row[0] == None:
                    break
                proxy_url = row[0]
                username = row[1]
                pwd = row[2]
                # proxy_address = str(username) + ':' + str(pwd) + '@' + str(proxy_url)
                proxy_address = {'url':proxy_url, 'username':username, 'pwd':pwd}
                self.proxies.append(proxy_address)
            self.proxies.pop(0)
        except Exception as e:
            print('There is error: ', e, 'reading proxies.') 

    def get_response(self, url):
        print('Sending Request with : ', url)
        try:
            i = random.randint(0, len(self.proxies) - 1)
            self.session.proxies = {
                str.format('http://{}', self.proxies[i]),
                str.format('https://{}', self.proxies[i]),
            }
            response = self.session.get(url)
            return response
        except Exception as e:
            print('There is error :', e,'sending request')

    def compare_closet(self, str1='', str2=''):
        tokens1 = word_tokenize(str1.lower())
        tokens2 = word_tokenize(str2.lower())
        sentences = [tokens1, tokens2]
        model = Word2Vec(sentences, min_count = 1)
        similarity = model.wv.n_similarity(tokens1, tokens2)
        
        if similarity > SIMILARITY_RATE:
            return similarity
        else:
            return 0

    def return_detailpage(self, url):
        
        self.page.goto(url)
        content = self.page.content()
        return content        

def main():
    extracotr = Extractor()
    extracotr.read_proxies(PROXIES)
    extracotr.get_all()

if __name__ == "__main__":
    main()
